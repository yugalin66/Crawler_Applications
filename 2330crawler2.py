# 抓取股票每天交易資料並儲存到2330.xlsx
import bs4
import urllib.request as req
import openpyxl as xl
from openpyxl import Workbook
from pathlib import Path
import time
import datetime

today = str(datetime.date.today())


############### can be modified ##################
stock_number = "2330"
sleep_time = 3
reference_day = 100  # 觀察前 n 天的股價
prediction_day = 3  # 預測 m 天後漲跌
start_year = 2021
end_year = 2023
##################################################

# create workbook if it doesn't exist, delete file if exists
filename = stock_number + ".xlsx"
path = Path(filename)
if path.exists():
    path.unlink()

wb = Workbook()
modified_sheet = wb.worksheets[0]
modified_sheet.title = "modified_data"
original_sheet = wb.create_sheet("original_data")
today_sheet = wb.create_sheet("today_data")


# set titles of original data sheet
def set_title(original_sheet, modified_sheet, today_sheet, reference_day):
    original_sheet.cell(1, 1).value = "date"
    original_sheet.cell(1, 2).value = "volume"
    original_sheet.cell(1, 3).value = "amount"
    original_sheet.cell(1, 4).value = "open"
    original_sheet.cell(1, 5).value = "high"
    original_sheet.cell(1, 6).value = "low"
    original_sheet.cell(1, 7).value = "close"
    original_sheet.cell(1, 8).value = "change"

    modified_sheet.cell(1, 1).value = "date"
    for i in range(2, reference_day+2):
        modified_sheet.cell(1, i).value = i - 1
        today_sheet.cell(1, i).value = i - 1
    modified_sheet.cell(1, reference_day + 2).value = "prediction"

    today_sheet.cell(1, 1).value = "date"


# write original data sheet
def WriteExcel(sheet, datas):
    row = sheet.max_row + 1

    for data in datas:
        for i in range(8):
            if data[i] == "X0.00":
                sheet.cell(row, i+1).value = "0"
            elif data[i][0] == "+":
                sheet.cell(row, i+1).value = data[i].replace("+", "")
            else:
                sheet.cell(row, i+1).value = data[i].replace(",", "")
        row += 1


# wirte modified data sheet
def ProcessExcel(original_sheet, modified_sheet, today_sheet, reference_day, prediction_day):
    row = modified_sheet.max_row + 1
    for day in range(reference_day + 1, original_sheet.max_row - prediction_day + 1):
        modified_sheet.cell(row, 1).value = original_sheet.cell(day, 1).value
        # rearrange close price of n reference days
        for pastdays in range(reference_day):
            modified_sheet.cell(
                row, pastdays + 2).value = original_sheet.cell(day - pastdays, 7).value
        # calculate change between today and n days after
        change = float(original_sheet.cell(day + prediction_day,
                       7).value) - float(original_sheet.cell(day, 7).value)
        if change < -20:
            section = "< -20"
        elif change >= -20 and change <= 0:
            section = "-20 ~ 0"
        elif change > 0 and change <= 20:
            section = "0 ~ 20"
        elif change > 20:
            section = "> 20"
        modified_sheet.cell(row, reference_day + 2).value = section
        row += 1

    today = original_sheet.max_row
    today_sheet.cell(2, 1).value = original_sheet.cell(today, 1).value
    for i in range(reference_day):
        today_sheet.cell(
            2, i + 2).value = original_sheet.cell(today - i, 7).value


# use web crawler to get stock information
def getData(url):
    request = req.Request(url, headers={
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"
    })
    with req.urlopen(request) as response:
        data = response.read().decode("utf-8")

    root = bs4.BeautifulSoup(data, "html.parser")

    datas = root.find_all("td")
    counter = 0
    info = []
    one_day_info = []
    for data in datas:
        if counter == 8:
            counter = 0
            info.append(one_day_info)
            one_day_info = []
            continue
        else:
            one_day_info.append(data.string)
            counter += 1

    return info


# main
set_title(original_sheet, modified_sheet, today_sheet, reference_day)
for year in range(start_year, end_year + 1):
    for month in range(1, 13):
        pageURL = "https://www.twse.com.tw/rwd/zh/afterTrading/STOCK_DAY?date=" + str(year) + \
            str(month).zfill(2) + "01&stockNo=" + \
            stock_number + "&response=html"  # 複製要抓取資料網頁的網址
        data = getData(pageURL)
        WriteExcel(original_sheet, data)
        time.sleep(sleep_time)

ProcessExcel(original_sheet, modified_sheet,
             today_sheet, reference_day, prediction_day)
wb.save(filename)
