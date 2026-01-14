# 抓取台積電前五天的交易量、總金額、開盤、最高、最低、收盤，並預測三天內漲跌
import bs4
import urllib.request as req
import openpyxl as xl
from openpyxl import Workbook
from pathlib import Path
import time

stock_number = "2330"
filename = stock_number + ".xlsx"

path = Path(filename)
if path.exists():
    wb = xl.load_workbook(filename)
    wb.remove_sheet(wb.worksheets[0])
    sheet = wb.create_sheet("Sheet")
else:
    wb = Workbook()
    sheet = wb.worksheets[0]


def set_title(sheet):
    sheet.cell(1, 1).value = "amount4"
    sheet.cell(1, 2).value = "money4"
    sheet.cell(1, 3).value = "open4"
    sheet.cell(1, 4).value = "high4"
    sheet.cell(1, 5).value = "low4"
    sheet.cell(1, 6).value = "close4"
    sheet.cell(1, 7).value = "change4"
    sheet.cell(1, 8).value = "amount3"
    sheet.cell(1, 9).value = "money3"
    sheet.cell(1, 10).value = "open3"
    sheet.cell(1, 11).value = "high3"
    sheet.cell(1, 12).value = "low3"
    sheet.cell(1, 13).value = "close3"
    sheet.cell(1, 14).value = "change3"
    sheet.cell(1, 15).value = "amount2"
    sheet.cell(1, 16).value = "money2"
    sheet.cell(1, 17).value = "open2"
    sheet.cell(1, 18).value = "high2"
    sheet.cell(1, 19).value = "low2"
    sheet.cell(1, 20).value = "close2"
    sheet.cell(1, 21).value = "change2"
    sheet.cell(1, 22).value = "amount1"
    sheet.cell(1, 23).value = "money1"
    sheet.cell(1, 24).value = "open1"
    sheet.cell(1, 25).value = "high1"
    sheet.cell(1, 26).value = "low1"
    sheet.cell(1, 27).value = "close1"
    sheet.cell(1, 28).value = "change1"
    sheet.cell(1, 29).value = "amount0"
    sheet.cell(1, 30).value = "money0"
    sheet.cell(1, 31).value = "open0"
    sheet.cell(1, 32).value = "high0"
    sheet.cell(1, 33).value = "low0"
    sheet.cell(1, 34).value = "close0"
    sheet.cell(1, 35).value = "change0"
    sheet.cell(1, 36).value = "result"


def WriteExcel(sheet, filename, datas):
    row = sheet.max_row + 1
    col = 1

    for i in range(len(datas)-8):
        for k in range(5):
            for j in range(len(datas[0])):
                if j == 0:
                    sheet.cell(row, col).value = datas[i+k][j][0:-8]
                elif j == 1:
                    sheet.cell(row, col).value = datas[i+k][j][0:-12]
                elif j == 6:
                    if datas[i+k][j] == "X0.00":
                        sheet.cell(row, col).value = 0.00
                    else:
                        sheet.cell(row, col).value = float(datas[i+k][j])
                else:
                    sheet.cell(row, col).value = datas[i+k][j]
                col += 1
            if k == 4:
                sum_in_3_days = 0
                for l in range(1, 4):
                    if datas[i+k+l][6] == "X0.00":
                        sum_in_3_days += 0
                    else:
                        sum_in_3_days += float(datas[i+k+l][6])

                if sum_in_3_days < -20.0:
                    result = "below -20"
                elif sum_in_3_days >= -20.0 and sum_in_3_days <= 0:
                    result = "-20 ~ 0"
                elif sum_in_3_days > 0 and sum_in_3_days < 20:
                    result = "0 ~ 20"
                elif sum_in_3_days >= 20:
                    result = "over 20"

                sheet.cell(row, col).value = result
                row += 1
                col = 1


def getData(url):
    # 建立一個Reqest物件，附加Request Headers的資訊
    # 按F12 -> Network，重新整理網頁 -> index.html -> Headers -> Request Headers ->複製需要的資訊
    request = req.Request(url, headers={
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"
    })
    with req.urlopen(request) as response:
        data = response.read().decode("utf-8")

    root = bs4.BeautifulSoup(data, "html.parser")

    datas = root.find_all("td")
    index = 0
    for data in datas:
        if index == 0:
            info = []
            index += 1
        elif index == 8:
            index = 0
        else:
            info.append(data.string)
            index += 1

    print(info)
    return info


# main : 抓取多個網頁的標題
set_title(sheet)
data = []
for year in range(2020, 2023):
    for month in range(1, 13):
        pageURL = "https://www.twse.com.tw/rwd/zh/afterTrading/STOCK_DAY?date=" + str(year) + \
            str(month).zfill(2) + "01&stockNo=" + \
            stock_number + "&response=html"  # 複製要抓取資料網頁的網址
        data.append(getData(pageURL))
        time.sleep(3)

print(data)
WriteExcel(sheet, filename, data)

wb.save(filename)
