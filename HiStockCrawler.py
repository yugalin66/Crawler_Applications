# 抓取PTT八卦版的網頁原始碼(HTML)
import bs4
import urllib.request as req
import openpyxl as xl
from openpyxl import Workbook
from pathlib import Path

filename = "stock.xlsx"
path = Path(filename)
if not path.exists():
    wb = Workbook()
    sheet = wb.active
    sheet.cell(1, 1).value = "編號"
    sheet.cell(1, 2).value = "昨收"
    sheet.cell(1, 3).value = "今開"
    wb.save(filename)
else:
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet"]
    sheet.delete_rows(1, sheet.max_row)


def WriteExcel(fname, no, open, yesterday):
    wb = xl.load_workbook(fname)
    sheet = wb["Sheet"]
    row = sheet.max_row + 1
    sheet.cell(row, 1).value = no
    sheet.cell(row, 2).value = yesterday
    sheet.cell(row, 3).value = open
    wb.save(fname)


def getData(url):
    # 建立一個Reqest物件，附加Request Headers的資訊
    # 按F12 -> Network，重新整理網頁 -> index.html -> Headers -> Request Headers ->複製需要的資訊
    request = req.Request(url, headers={
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"
    })
    with req.urlopen(request) as response:
        data = response.read().decode("utf-8")

    # 解析原始碼，取得每篇文章的標題
    root = bs4.BeautifulSoup(data, "html.parser")
    titles = root.find_all("td")  # 尋找所有 class="title" 的div標籤
    counter = 0
    for title in titles:
        if counter == 0:
            no = title.string
        elif counter == 7:
            open = title.string
        elif counter == 10:
            yesterday = title.string
        if counter == 12:
            WriteExcel(filename, no, open, yesterday)
            counter = 0
        else:
            counter += 1


# main : 抓取多個網頁的標題
pageURL = "https://histock.tw/stock/rank.aspx?p=all"  # 複製要抓取資料網頁的網址
getData(pageURL)
