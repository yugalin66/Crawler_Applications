# 抓每支股票的代碼、昨收、今開，並寫入 stock.xlsx
import bs4
import urllib.request as req
import openpyxl as xl
from openpyxl import Workbook
from pathlib import Path

date = "0815"
filename = "stock.xlsx"
pageURL = "https://histock.tw/stock/rank.aspx?p=all"
path = Path(filename)
if path.exists():
    wb = xl.load_workbook(filename)
    if date not in wb.sheetnames:
        sheet = wb.create_sheet(title=date)
    else:
        print("Already done!")
        exit(0)
else:
    wb = Workbook()
    sheet = wb.worksheets[0]
    sheet.title = date
sheet.cell(1, 1).value = "no"
sheet.cell(1, 2).value = "week change ratio"
sheet.cell(1, 3).value = "last close"
sheet.cell(1, 4).value = "today open"
sheet.cell(1, 5).value = "today close"
sheet.cell(1, 6).value = "change ratio"
sheet.cell(1, 7).value = "profitable"


def WriteExcel(sheet, infos):
    row = sheet.max_row + 1
    for info in infos:
        sheet.cell(row, 1).value = info[0]
        sheet.cell(row, 2).value = info[1]
        sheet.cell(row, 3).value = info[2]
        sheet.cell(row, 4).value = info[3]
        sheet.cell(row, 5).value = info[4]
        sheet.cell(row, 6).value = info[5]
        if info[5][0] == "-":
            sheet.cell(row, 7).value = 0
        else:
            sheet.cell(row, 7).value = 1
        row += 1


def getData(url):
    # 建立一個Reqest物件，附加Request Headers的資訊
    # 按F12 -> Network，重新整理網頁 -> index.html -> Headers -> Request Headers ->複製需要的資訊
    request = req.Request(url, headers={
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"
    })
    with req.urlopen(request) as response:
        data = response.read().decode("utf-8")

    # 解析原始碼，取得每篇文章的標題
    root = bs4.BeautifulSoup(data, "html.parser")
    titles = root.find_all("td")
    counter = 0
    info = []
    for title in titles:
        if counter == 0:
            no = title.string
        elif counter == 2:
            close = title.span.string
        elif counter == 4:
            change = title.span.string
            if change == "--":
                change = "0"
            change = change.replace("%", "")
            change = change.replace("+", "")
        elif counter == 5:
            week_change = title.span.string
            week_change = week_change.replace("%", "")
        elif counter == 7:
            open = title.string
        elif counter == 10:
            yesterday = title.string
        if counter == 12:
            if no.isdigit():
                info.append([no, week_change, yesterday, open, close, change])
            counter = 0
        else:
            counter += 1

    return info


# main : 抓取多個網頁的標題

infos = getData(pageURL)
WriteExcel(sheet, infos)
wb.save(filename)
